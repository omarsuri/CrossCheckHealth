import json
import math
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl


INPUT_XLSX = Path(r"C:\Users\faroo\Downloads\indian_health_products_with_image_links_pass1.xlsx")
OUTPUT_SQL = Path(__file__).resolve().parents[1] / "supabase" / "import-indian-health-products-with-images.sql"


def clean(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    return text if text and text.lower() != "nan" else None


def sql_text(value):
    value = clean(value)
    if value is None:
        return "null"
    return "'" + value.replace("'", "''") + "'"


def sql_date(value):
    value = clean(value)
    if value is None:
        return "null"
    return sql_text(value[:10])


def slugify(value):
    value = (clean(value) or "product").lower()
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return value or "product"


def stable_id(row):
    return "IMG-" + "-".join(
        part for part in [
            slugify(row.get("Category")),
            slugify(row.get("Brand")),
            slugify(row.get("Product Name")),
        ] if part
    )[:140]


def is_direct_image(url, status):
    url = clean(url) or ""
    status = (clean(status) or "").lower()
    if not url.startswith("http"):
        return False
    if "google.com/search" in url or "tbm=isch" in url:
        return False
    return "direct" in status or re.search(r"\.(png|jpg|jpeg|webp)(\?|$)", url, re.I) is not None


def verdict_for(row):
    review = (clean(row.get("Review Status")) or "").lower()
    image_status = (clean(row.get("Image Status")) or "").lower()
    if "product-level" in review and "direct" in image_status:
        return "Strong Choice", "yes", "Product-level source and direct official image were verified in the workbook."
    return "Compare First", "conditional", "Source/category was captured, but this row should be manually checked before publishing claims."


def scores_for(row):
    review = (clean(row.get("Review Status")) or "").lower()
    image_status = (clean(row.get("Image Status")) or "").lower()
    has_source = bool(clean(row.get("Source URL")))
    direct_image = "direct" in image_status
    product_level = "product-level" in review
    transparency = 88 if product_level and direct_image else 72 if has_source else 45
    science = 50
    value = 0
    safety = 55
    efficacy = 50
    family = 55
    overall = round((science + safety + efficacy + transparency + family) / 5)
    return science, value, safety, family, transparency, efficacy, overall


def read_products():
    workbook = openpyxl.load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    sheet = workbook["Products"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(clean(value) for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def main():
    rows = read_products()
    category_pairs = sorted({(clean(row.get("Category")), slugify(row.get("Category"))) for row in rows if clean(row.get("Category"))})
    product_rows = []
    score_rows = []
    ingredient_rows = []
    warning_rows = []
    tag_rows = []

    for row in rows:
        source_product_id = stable_id(row)
        category = clean(row.get("Category"))
        category_slug = slugify(category)
        name = clean(row.get("Product Name"))
        brand = clean(row.get("Brand"))
        product_type = clean(row.get("Product Type"))
        primary_use = clean(row.get("Primary Use / Positioning"))
        form = clean(row.get("Form"))
        pack_variant = clean(row.get("Pack / Variant"))
        specs = clean(row.get("Key Ingredients / Specs"))
        source_name = clean(row.get("Source Name"))
        source_url = clean(row.get("Source URL"))
        product_image_link = clean(row.get("Product Image Link"))
        image_status = clean(row.get("Image Status"))
        direct_image = is_direct_image(product_image_link, image_status)
        image_url = product_image_link if direct_image else None
        image_search_url = None if direct_image else product_image_link
        review_status = clean(row.get("Review Status"))
        verdict, verdict_key, interpretation = verdict_for(row)
        practical_take = primary_use or product_type or "Product catalogue entry for preventive healthcare browsing."
        image_note = clean(row.get("Image Verification Notes"))
        notes = clean(row.get("Notes"))
        availability = clean(row.get("India Availability Evidence"))
        image_link_type = clean(row.get("Image Link Type"))
        last_checked = clean(row.get("Last Checked"))
        chips = [item for item in [
            product_type,
            form,
            "Direct Image" if direct_image else "Image Needs Review",
            "Source Verified" if "product-level" in (review_status or "").lower() else "Manual Check",
        ] if item]
        certifications = [source_name] if source_name else []
        consumer_transparency = " | ".join(item for item in [
            f"Source: {source_name}" if source_name else None,
            f"URL: {source_url}" if source_url else None,
            review_status,
            availability,
        ] if item)
        exposure_interpretation = " | ".join(item for item in [
            image_status,
            image_link_type,
            image_note,
            notes,
        ] if item)

        product_rows.append(
            "(" + ", ".join([
                sql_text(source_product_id),
                sql_text(name),
                sql_text(brand),
                f"(select id from product_categories where slug={sql_text(category_slug)})",
                sql_text(category),
                sql_text(category_slug),
                sql_text(product_type),
                sql_text(product_type),
                sql_text(primary_use),
                sql_text(form),
                sql_text(pack_variant),
                "null",
                "null",
                sql_text(pack_variant),
                "'INR'",
                "null",
                "0",
                sql_text(verdict),
                sql_text(interpretation),
                sql_text(practical_take),
                sql_text(name),
                sql_text(practical_take),
                sql_text(source_url),
                sql_text(image_url),
                sql_text(image_search_url),
                sql_text(image_status),
                sql_text(image_link_type),
                sql_text(image_note),
                sql_text(source_name),
                sql_text(source_url),
                sql_text(availability),
                sql_text(review_status),
                sql_text(notes),
                sql_date(last_checked),
                sql_text(json.dumps(certifications)),
                sql_text(json.dumps(chips)),
                sql_text(specs or "Ingredient/spec details are not listed in this workbook row."),
                sql_text(consumer_transparency or "Source evidence not listed."),
                sql_text(primary_use or "Preventive product catalogue entry only; not medical advice."),
                sql_text(exposure_interpretation or "Image/source audit notes are not listed."),
                sql_text(review_status or "Source review not listed."),
                "true",
            ]) + ")"
        )

        science, value, safety, family, transparency, efficacy, overall = scores_for(row)
        score_rows.append(
            f"({sql_text(source_product_id)}, {science}, {value}, {safety}, {family}, {family}, {transparency}, {efficacy}, 20, {overall}, {sql_text(verdict)}, {sql_text(verdict_key)}, {sql_text(interpretation)})"
        )

        ingredient_name = specs or product_type or form or "Specs not listed"
        ingredient_rows.append(
            f"({sql_text(source_product_id)}, {sql_text(ingredient_name)}, {sql_text(pack_variant or 'Not listed')}, 'und', 'not assessed', {sql_text(product_type or 'Product spec')}, {sql_text('Workbook listed ingredient/spec: ' + ingredient_name)})"
        )

        if not direct_image:
            warning_rows.append(
                f"({sql_text(source_product_id)}, 'Image requires manual review', {sql_text('The workbook provides an image-search link, not a verified direct product image URL.')}, 'medium', 'Image review')"
            )
        if "manual cross-check" in (review_status or "").lower() or "category verified" in (review_status or "").lower():
            warning_rows.append(
                f"({sql_text(source_product_id)}, 'Product row needs manual cross-check', {sql_text(review_status)}, 'medium', 'Verify source')"
            )

        for chip in chips:
            tag_rows.append(f"({sql_text(source_product_id)}, {sql_text(chip)}, 'workbook')")

    product_columns = "source_product_id, name, brand, category_id, category, cat, subcategory, product_type, primary_use, form, pack_variant, price, original_price, price_unit, currency, rating, review_count, verdict, interpretation, practical_take, usp_headline, usp_context, affiliate_url, image_url, image_search_url, image_status, image_link_type, image_verification_notes, source_name, source_url, india_availability_evidence, review_status, source_notes, last_checked, certifications, chips, ingredient_research, consumer_transparency, wellness_context, exposure_interpretation, evidence_strength, is_active"
    product_update = "name=excluded.name, brand=excluded.brand, category_id=excluded.category_id, category=excluded.category, cat=excluded.cat, subcategory=excluded.subcategory, product_type=excluded.product_type, primary_use=excluded.primary_use, form=excluded.form, pack_variant=excluded.pack_variant, price_unit=excluded.price_unit, verdict=excluded.verdict, interpretation=excluded.interpretation, practical_take=excluded.practical_take, usp_headline=excluded.usp_headline, usp_context=excluded.usp_context, affiliate_url=excluded.affiliate_url, image_url=excluded.image_url, image_search_url=excluded.image_search_url, image_status=excluded.image_status, image_link_type=excluded.image_link_type, image_verification_notes=excluded.image_verification_notes, source_name=excluded.source_name, source_url=excluded.source_url, india_availability_evidence=excluded.india_availability_evidence, review_status=excluded.review_status, source_notes=excluded.source_notes, last_checked=excluded.last_checked, certifications=excluded.certifications, chips=excluded.chips, ingredient_research=excluded.ingredient_research, consumer_transparency=excluded.consumer_transparency, wellness_context=excluded.wellness_context, exposure_interpretation=excluded.exposure_interpretation, evidence_strength=excluded.evidence_strength, is_active=excluded.is_active, updated_at=now()"

    lines = [
        "-- Auto-generated from indian_health_products_with_image_links_pass1.xlsx",
        "-- Contains 220 product catalogue rows. Direct product images are stored in image_url.",
        "-- Image-search links are stored in image_search_url for manual review and are not used as card images.",
        "",
        "alter table products add column if not exists source_product_id text;",
        "alter table products add column if not exists product_type text;",
        "alter table products add column if not exists primary_use text;",
        "alter table products add column if not exists pack_variant text;",
        "alter table products add column if not exists image_search_url text;",
        "alter table products add column if not exists image_status text;",
        "alter table products add column if not exists image_link_type text;",
        "alter table products add column if not exists image_verification_notes text;",
        "alter table products add column if not exists source_name text;",
        "alter table products add column if not exists source_url text;",
        "alter table products add column if not exists india_availability_evidence text;",
        "alter table products add column if not exists review_status text;",
        "alter table products add column if not exists source_notes text;",
        "alter table products add column if not exists last_checked date;",
        "alter table products add column if not exists certifications jsonb default '[]'::jsonb;",
        "alter table products add column if not exists chips jsonb default '[]'::jsonb;",
        "alter table products add column if not exists ingredient_research text;",
        "alter table products add column if not exists consumer_transparency text;",
        "alter table products add column if not exists wellness_context text;",
        "alter table products add column if not exists exposure_interpretation text;",
        "alter table products add column if not exists evidence_strength text;",
        "alter table product_scores add column if not exists family_safety_score int default 0;",
        "alter table product_ingredients add column if not exists ingredient_type text;",
        "alter table product_ingredients add column if not exists microcopy text;",
        "alter table product_warnings add column if not exists caution_label text;",
        "create table if not exists product_tags (",
        "  id uuid primary key default gen_random_uuid(),",
        "  product_id uuid references products(id) on delete cascade,",
        "  label text not null,",
        "  tag_type text default 'tag',",
        "  created_at timestamptz default now(),",
        "  unique(product_id, label)",
        ");",
        "drop index if exists products_source_product_id_key;",
        "create unique index if not exists products_source_product_id_key on products(source_product_id);",
        "",
        "insert into product_categories (name, slug) values",
        ",\n".join(f"  ({sql_text(name)}, {sql_text(slug)})" for name, slug in category_pairs) + "\non conflict (slug) do nothing;",
        "",
        "-- Products are emitted as complete per-row upserts so partial execution does not leave a dangling comma.",
        "\n".join(f"insert into products ({product_columns}) values\n  {product_row}\non conflict (source_product_id) do update set\n  {product_update};" for product_row in product_rows),
        "",
        "delete from product_scores where product_id in (select id from products where source_product_id like 'IMG-%');",
        "insert into product_scores (product_id, science_score, value_score, safety_score, parent_score, family_safety_score, transparency_score, efficacy_score, hype_score, overall_score, verdict, verdict_key, verdict_text)",
        "select p.id, v.science_score, v.value_score, v.safety_score, v.parent_score, v.family_safety_score, v.transparency_score, v.efficacy_score, v.hype_score, v.overall_score, v.verdict, v.verdict_key, v.verdict_text from (values",
        ",\n".join("  " + score_row for score_row in score_rows),
        ") as v(source_product_id, science_score, value_score, safety_score, parent_score, family_safety_score, transparency_score, efficacy_score, hype_score, overall_score, verdict, verdict_key, verdict_text) join products p on p.source_product_id=v.source_product_id;",
        "",
        "delete from product_ingredients where product_id in (select id from products where source_product_id like 'IMG-%');",
        "insert into product_ingredients (product_id, ingredient_name, amount, status, evidence_level, ingredient_type, microcopy)",
        "select p.id, v.ingredient_name, v.amount, v.status, v.evidence_level, v.ingredient_type, v.microcopy from (values",
        ",\n".join("  " + ingredient_row for ingredient_row in ingredient_rows),
        ") as v(source_product_id, ingredient_name, amount, status, evidence_level, ingredient_type, microcopy) join products p on p.source_product_id=v.source_product_id;",
        "",
        "delete from product_warnings where product_id in (select id from products where source_product_id like 'IMG-%');",
    ]
    if warning_rows:
        lines.extend([
            "insert into product_warnings (product_id, warning_title, warning_text, severity, caution_label)",
            "select p.id, v.warning_title, v.warning_text, v.severity, v.caution_label from (values",
            ",\n".join("  " + warning_row for warning_row in warning_rows),
            ") as v(source_product_id, warning_title, warning_text, severity, caution_label) join products p on p.source_product_id=v.source_product_id;",
        ])
    lines.extend([
        "",
        "delete from product_tags where product_id in (select id from products where source_product_id like 'IMG-%');",
        "insert into product_tags (product_id, label, tag_type)",
        "select p.id, v.label, v.tag_type from (values",
        ",\n".join("  " + tag_row for tag_row in tag_rows),
        ") as v(source_product_id, label, tag_type) join products p on p.source_product_id=v.source_product_id on conflict (product_id, label) do nothing;",
        "",
    ])
    OUTPUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    direct_count = sum(1 for row in rows if is_direct_image(row.get("Product Image Link"), row.get("Image Status")))
    print(f"Wrote {OUTPUT_SQL}")
    print(f"Products: {len(rows)}")
    print(f"Direct images: {direct_count}")
    print(f"Image review links: {len(rows) - direct_count}")


if __name__ == "__main__":
    main()
