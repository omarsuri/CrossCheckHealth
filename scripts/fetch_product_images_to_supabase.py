import argparse
import csv
import html
import json
import mimetypes
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode, urljoin, urlparse, urlunparse
from urllib.request import Request, urlopen
from urllib.robotparser import RobotFileParser

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = ROOT / ".env.local"
BUCKET = "product-images"
MIN_IMAGE_BYTES = 4_000
MAX_IMAGE_BYTES = 4_500_000
MIN_IMAGE_SIDE = 180
PREFERRED_IMAGE_SIDE = 300
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 CrossCheckHealthImageBot/1.0"
TRUSTED_RETAILER_DOMAINS = {
    "tata1mg.com",
    "www.tata1mg.com",
    "1mg.com",
    "www.1mg.com",
    "apollopharmacy.in",
    "www.apollopharmacy.in",
    "netmeds.com",
    "www.netmeds.com",
    "pharmeasy.in",
    "www.pharmeasy.in",
    "healthkart.com",
    "www.healthkart.com",
}
GOOGLE_HOSTS = {"google.com", "www.google.com", "images.google.com"}


@dataclass
class Candidate:
    url: str
    source_page: str
    source_domain: str
    confidence_score: float
    width: int
    height: int
    content_type: str
    extension: str
    data: bytes


class Summary:
    def __init__(self):
        self.total_products_checked = 0
        self.skipped_existing_images = 0
        self.skipped_invalid_source_urls = 0
        self.blocked_by_robots = 0
        self.source_pages_fetched = 0
        self.candidates_found = 0
        self.images_uploaded = 0
        self.manual_required = 0
        self.failed = 0

    def print(self):
        print("\nSummary:")
        for key, value in self.__dict__.items():
            print(f"- {key}: {value}")


class ImageCandidateParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.meta_images: list[tuple[str, str, int]] = []
        self.image_src_links: list[tuple[str, str, int]] = []
        self.imgs: list[dict[str, str]] = []
        self.json_ld: list[str] = []
        self._in_json_ld = False
        self._json_ld_buffer: list[str] = []

    def handle_starttag(self, tag, attrs):
        attr = {key.lower(): html.unescape(value or "") for key, value in attrs}
        tag = tag.lower()
        if tag == "meta":
            key = (attr.get("property") or attr.get("name") or "").lower()
            if key in {"og:image", "og:image:url", "twitter:image", "twitter:image:src"} and attr.get("content"):
                self.meta_images.append((attr["content"], key, 0))
        elif tag == "link":
            rel = attr.get("rel", "").lower()
            if "image_src" in rel and attr.get("href"):
                self.image_src_links.append((attr["href"], rel, 0))
        elif tag == "img":
            src = attr.get("src") or attr.get("data-src") or attr.get("data-lazy-src") or attr.get("data-original")
            srcset = attr.get("srcset") or attr.get("data-srcset")
            if src or srcset:
                self.imgs.append({
                    "src": src or "",
                    "srcset": srcset or "",
                    "alt": attr.get("alt", ""),
                    "class": attr.get("class", ""),
                    "id": attr.get("id", ""),
                    "width": attr.get("width", ""),
                    "height": attr.get("height", ""),
                })
        elif tag == "script" and "ld+json" in attr.get("type", "").lower():
            self._in_json_ld = True
            self._json_ld_buffer = []

    def handle_data(self, data):
        if self._in_json_ld:
            self._json_ld_buffer.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "script" and self._in_json_ld:
            self.json_ld.append("".join(self._json_ld_buffer))
            self._in_json_ld = False
            self._json_ld_buffer = []


def load_env(path: Path):
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def request_bytes(url: str, method="GET", headers=None, body=None, timeout=22):
    request = Request(url, data=body, method=method, headers={
        "User-Agent": USER_AGENT,
        **(headers or {}),
    })
    with urlopen(request, timeout=timeout) as response:
        return response.status, dict(response.headers), response.read()


def request_json(url: str, method="GET", headers=None, body=None, timeout=35):
    status, _headers, data = request_bytes(url, method=method, headers=headers, body=body, timeout=timeout)
    if status < 200 or status >= 300:
        raise RuntimeError(f"HTTP {status}: {url}")
    if not data:
        return None
    return json.loads(data.decode("utf-8"))


def supabase_headers(service_key: str, content_type="application/json"):
    return {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": content_type,
    }


def normalize_header(value: str):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text.lower() not in {"nan", "none", "null"} else None


def detect_sheet(default_sheet: str | None):
    if default_sheet:
        return Path(default_sheet)
    candidates = sorted([*ROOT.glob("*.xlsx"), *ROOT.glob("*.csv")], key=lambda path: path.stat().st_mtime, reverse=True)
    if not candidates:
        return None
    return candidates[0]


def read_sheet(path: Path | None):
    if not path:
        return []
    if not path.exists():
        raise SystemExit(f"Sheet not found: {path}")
    rows = []
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                normalized = {normalize_header(key): clean(value) for key, value in row.items()}
                if any(normalized.values()):
                    rows.append(normalized)
    elif path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["Products"] if "Products" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        headers = [normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers)}
            if any(row.values()):
                rows.append(row)
    else:
        raise SystemExit("Only .xlsx and .csv sheets are supported")
    return rows


def normalize_key(*parts):
    return re.sub(r"[^a-z0-9]+", "", " ".join(clean(part) or "" for part in parts).lower())


def sheet_product_id(row):
    for key in ["product_id", "source_product_id", "id"]:
        if row.get(key):
            return row[key]
    return None


def get_first(row, keys):
    for key in keys:
        if row.get(key):
            return row[key]
    return None


def build_sheet_indexes(rows):
    by_id = {}
    by_brand_name = {}
    for row in rows:
        product_id = sheet_product_id(row)
        if product_id:
            by_id[str(product_id)] = row
        brand = get_first(row, ["brand", "brand_name"])
        name = get_first(row, ["product_name", "name", "product"])
        if brand and name:
            by_brand_name[normalize_key(brand, name)] = row
    return by_id, by_brand_name


def fetch_products(supabase_url: str, service_key: str, only_missing: bool, force: bool, limit: int):
    query = {
        "select": "id,source_product_id,name,brand,category,source_url,image_url,image_search_url,image_status",
        "is_active": "eq.true",
        "order": "created_at.asc",
        "limit": str(limit),
    }
    if only_missing and not force:
        query["image_url"] = "is.null"
    url = f"{supabase_url.rstrip('/')}/rest/v1/products?{urlencode(query)}"
    return request_json(url, headers=supabase_headers(service_key)) or []


def is_google_url(url):
    parsed = urlparse(url or "")
    host = parsed.netloc.lower().removeprefix("www.")
    return host in {host.removeprefix("www.") for host in GOOGLE_HOSTS} or "google.com/search" in (url or "")


def is_valid_source_url(url):
    if not url or not url.startswith(("http://", "https://")):
        return False
    if is_google_url(url):
        return False
    path = urlparse(url).path.lower()
    if re.search(r"\.(jpg|jpeg|png|webp|gif|svg|ico)(\?|$)", path):
        return False
    return True


def source_kind(url):
    domain = urlparse(url).netloc.lower()
    if domain in TRUSTED_RETAILER_DOMAINS or domain.removeprefix("www.") in TRUSTED_RETAILER_DOMAINS:
        return "retailer"
    return "official"


def choose_source_urls(product, sheet_row, source_mode):
    candidates = [
        product.get("source_url"),
        get_first(sheet_row or {}, ["official_url", "official_product_url", "brand_url"]),
        get_first(sheet_row or {}, ["retailer_url", "retailer_product_url", "product_url"]),
        get_first(sheet_row or {}, ["source_url", "source_link"]),
    ]
    urls = []
    for url in candidates:
        if not is_valid_source_url(url):
            continue
        kind = source_kind(url)
        if source_mode != "all" and kind != source_mode:
            continue
        if url not in urls:
            urls.append(url)
    return urls


def safe_url(url: str):
    parsed = urlparse(url.strip())
    path = quote(parsed.path, safe="/:%")
    query = quote(parsed.query, safe="=&?/:+,%")
    return urlunparse((parsed.scheme, parsed.netloc, path, parsed.params, query, parsed.fragment))


def robots_allowed(url, robots_cache):
    parsed = urlparse(url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    if base not in robots_cache:
        parser = RobotFileParser()
        parser.set_url(urljoin(base, "/robots.txt"))
        try:
            parser.read()
        except Exception:
            parser = None
        robots_cache[base] = parser
    parser = robots_cache[base]
    if parser is None:
        return True
    try:
        return parser.can_fetch(USER_AGENT, url)
    except Exception:
        return True


def largest_from_srcset(srcset: str | None):
    if not srcset:
        return None
    candidates = []
    for part in srcset.split(","):
        pieces = part.strip().split()
        if not pieces:
            continue
        url = pieces[0]
        weight = 0
        if len(pieces) > 1:
            match = re.match(r"(\d+)(w|x)", pieces[1])
            if match:
                weight = int(match.group(1))
        candidates.append((weight, url))
    return sorted(candidates)[-1][1] if candidates else None


def extract_json_images(value):
    images = []
    if isinstance(value, str) and value.startswith(("http://", "https://")):
        images.append(value)
    elif isinstance(value, list):
        for item in value:
            images.extend(extract_json_images(item))
    elif isinstance(value, dict):
        type_value = value.get("@type")
        if isinstance(type_value, list):
            is_product = any(str(item).lower() == "product" for item in type_value)
        else:
            is_product = str(type_value or "").lower() == "product"
        if is_product and "image" in value:
            images.extend(extract_json_images(value["image"]))
        for key in ["image", "thumbnailUrl", "contentUrl"]:
            if key in value:
                images.extend(extract_json_images(value[key]))
    return images


def tokenize(*parts) -> set[str]:
    text = " ".join(part or "" for part in parts).lower()
    tokens = {token for token in re.split(r"[^a-z0-9]+", text) if len(token) >= 3}
    return tokens - {"the", "and", "for", "with", "india", "product", "tablet", "capsule", "capsules", "health"}


def raw_candidate_score(url, context, product, priority):
    lower = f"{url} {context}".lower()
    if lower.startswith("data:"):
        return -100
    reject_words = ["logo", "icon", "sprite", "placeholder", "banner", "payment", "avatar", "favicon", "loader", "pixel", "tracking"]
    score = priority
    if any(word in lower for word in reject_words):
        score -= 80
    tokens = tokenize(product.get("name"), product.get("brand"))
    score += min(40, sum(8 for token in tokens if token in lower))
    if any(word in lower for word in ["product", "pack", "front", "gallery", "media", "catalog", "cdn", "images"]):
        score += 12
    if any(domain in lower for domain in ["1mg", "apollo", "netmeds", "pharmeasy", "healthkart"]):
        score += 8
    return score


def extract_image_urls(page_url, html_text, product):
    parser = ImageCandidateParser()
    parser.feed(html_text)
    found = []

    def add(url, context="", priority=0):
        if not url:
            return
        absolute = safe_url(urljoin(page_url, html.unescape(url)))
        if is_google_url(absolute) or absolute.startswith("data:"):
            return
        found.append((raw_candidate_score(absolute, context, product, priority), absolute, context))

    for url, context, _ in parser.meta_images:
        add(url, context, 90)
    for url, context, _ in parser.image_src_links:
        add(url, context, 70)
    for blob in parser.json_ld:
        try:
            data = json.loads(blob)
        except Exception:
            continue
        for url in extract_json_images(data):
            add(url, "json-ld product image", 85)
    for image in parser.imgs:
        src = largest_from_srcset(image.get("srcset")) or image.get("src")
        context = " ".join([image.get("alt") or "", image.get("class") or "", image.get("id") or ""])
        add(src, context, 40)

    deduped = {}
    for score, url, context in found:
        if url not in deduped or score > deduped[url][0]:
            deduped[url] = (score, context)
    return [(url, score, context) for url, (score, context) in sorted(deduped.items(), key=lambda item: item[1][0], reverse=True)]


def validate_image(url):
    status, headers, data = request_bytes(url, headers={"Accept": "image/avif,image/webp,image/png,image/jpeg,image/*,*/*"})
    content_type = headers.get("Content-Type", "").split(";")[0].lower()
    if status < 200 or status >= 300:
        return None
    if len(data) < MIN_IMAGE_BYTES or len(data) > MAX_IMAGE_BYTES:
        return None
    if not content_type.startswith("image/"):
        guessed, _ = mimetypes.guess_type(urlparse(url).path)
        if not guessed or not guessed.startswith("image/"):
            return None
        content_type = guessed
    try:
        image = Image.open(BytesIO(data))
        width, height = image.size
        image.verify()
    except Exception:
        return None
    if width < MIN_IMAGE_SIDE or height < MIN_IMAGE_SIDE:
        return None
    extension = mimetypes.guess_extension(content_type) or ".jpg"
    if extension == ".jpe":
        extension = ".jpg"
    return {"data": data, "content_type": content_type, "extension": extension, "width": width, "height": height}


def confidence(raw_score, width, height, url, source_page):
    score = max(0, min(100, raw_score))
    if width >= PREFERRED_IMAGE_SIDE and height >= PREFERRED_IMAGE_SIDE:
        score += 12
    if abs(width - height) <= max(width, height) * 0.25:
        score += 8
    if urlparse(url).netloc == urlparse(source_page).netloc:
        score += 6
    if source_kind(source_page) == "retailer":
        score += 4
    return round(max(0, min(1, score / 100)), 3)


def fetch_page(url):
    status, headers, data = request_bytes(url, headers={"Accept": "text/html,application/xhtml+xml"})
    content_type = headers.get("Content-Type", "")
    if status < 200 or status >= 300 or "text/html" not in content_type:
        return None
    return data.decode("utf-8", errors="ignore")


def slugify(value: str):
    value = (value or "product").lower()
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return value[:80] or "product"


def ensure_bucket(supabase_url: str, service_key: str, dry_run: bool):
    if dry_run:
        return
    url = f"{supabase_url.rstrip('/')}/storage/v1/bucket"
    body = json.dumps({"id": BUCKET, "name": BUCKET, "public": True, "file_size_limit": MAX_IMAGE_BYTES}).encode("utf-8")
    try:
        request_json(url, method="POST", headers=supabase_headers(service_key), body=body)
        print(f"Created storage bucket: {BUCKET}")
    except HTTPError as error:
        if error.code not in {400, 409}:
            raise


def storage_public_url(supabase_url, object_path):
    return f"{supabase_url.rstrip('/')}/storage/v1/object/public/{BUCKET}/{object_path}"


def upload_image(supabase_url, service_key, product, candidate: Candidate, dry_run):
    name = slugify(product.get("name") or product["id"])
    extension = ".jpg" if candidate.extension.lower() in {".jpe", ".jpeg"} else candidate.extension
    object_path = f"{product['id']}/{name}{extension}"
    if dry_run:
        return storage_public_url(supabase_url, object_path)
    url = f"{supabase_url.rstrip('/')}/storage/v1/object/{BUCKET}/{quote(object_path)}"
    headers = supabase_headers(service_key, candidate.content_type)
    headers["x-upsert"] = "true"
    request_bytes(url, method="POST", headers=headers, body=candidate.data, timeout=40)
    return storage_public_url(supabase_url, object_path)


def patch_product(supabase_url, service_key, product_id, fields, dry_run):
    if dry_run:
        return
    url = f"{supabase_url.rstrip('/')}/rest/v1/products?id=eq.{product_id}"
    headers = supabase_headers(service_key)
    headers["Prefer"] = "return=minimal"
    request_bytes(url, method="PATCH", headers=headers, body=json.dumps(fields).encode("utf-8"))


def save_candidates(supabase_url, service_key, product_id, candidates: list[Candidate], approved_url, dry_run):
    if dry_run or not candidates:
        return
    rows = []
    for candidate in candidates:
        rows.append({
            "product_id": product_id,
            "candidate_url": candidate.url,
            "source_page": candidate.source_page,
            "source_domain": candidate.source_domain,
            "confidence_score": candidate.confidence_score,
            "width": candidate.width,
            "height": candidate.height,
            "status": "approved" if candidate.url == approved_url else "pending",
        })
    url = f"{supabase_url.rstrip('/')}/rest/v1/product_image_candidates?on_conflict=product_id,candidate_url"
    headers = supabase_headers(service_key)
    headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
    request_bytes(url, method="POST", headers=headers, body=json.dumps(rows).encode("utf-8"))


def mark_status(supabase_url, service_key, product_id, status, dry_run):
    patch_product(supabase_url, service_key, product_id, {
        "image_status": status,
        "image_checked_at": datetime.now(timezone.utc).isoformat(),
    }, dry_run)


def process_product(product, sheet_row, args, supabase_url, service_key, robots_cache, summary: Summary):
    summary.total_products_checked += 1
    label = f"{product.get('brand') or ''} {product.get('name') or product.get('id')}".strip()
    if product.get("image_url") and args.only_missing and not args.force:
        summary.skipped_existing_images += 1
        print(f"SKIP existing image: {label}")
        return

    source_urls = choose_source_urls(product, sheet_row, args.source)
    if not source_urls:
        summary.skipped_invalid_source_urls += 1
        mark_status(supabase_url, service_key, product["id"], "manual_required", args.dry_run)
        print(f"SKIP no valid source page: {label}")
        return

    valid_candidates = []
    for source_url in source_urls:
        if not robots_allowed(source_url, robots_cache):
            summary.blocked_by_robots += 1
            mark_status(supabase_url, service_key, product["id"], "blocked_by_robots", args.dry_run)
            print(f"BLOCKED by robots.txt: {label} - {source_url}")
            continue
        try:
            page = fetch_page(source_url)
        except Exception as error:
            print(f"FETCH FAILED: {label} - {source_url} - {error}")
            continue
        if not page:
            continue
        summary.source_pages_fetched += 1
        raw_candidates = extract_image_urls(source_url, page, product)
        for image_url, raw_score, _context in raw_candidates[:args.max_candidates]:
            try:
                image = validate_image(image_url)
            except Exception:
                image = None
            if not image:
                continue
            candidate = Candidate(
                url=image_url,
                source_page=source_url,
                source_domain=urlparse(source_url).netloc.lower(),
                confidence_score=confidence(raw_score, image["width"], image["height"], image_url, source_url),
                width=image["width"],
                height=image["height"],
                content_type=image["content_type"],
                extension=image["extension"],
                data=image["data"],
            )
            valid_candidates.append(candidate)
        time.sleep(args.delay)

    valid_candidates.sort(key=lambda candidate: candidate.confidence_score, reverse=True)
    summary.candidates_found += len(valid_candidates)

    if not valid_candidates:
        summary.failed += 1
        mark_status(supabase_url, service_key, product["id"], "manual_required", args.dry_run)
        print(f"FAILED no valid image candidates: {label}")
        return

    best = valid_candidates[0]
    approved_url = None
    if best.confidence_score >= args.min_confidence:
        public_url = upload_image(supabase_url, service_key, product, best, args.dry_run)
        patch_product(supabase_url, service_key, product["id"], {
            "image_url": public_url,
            "image_source_url": best.url,
            "image_status": "found",
            "image_checked_at": datetime.now(timezone.utc).isoformat(),
        }, args.dry_run)
        approved_url = best.url
        summary.images_uploaded += 1
        print(f"FOUND {best.confidence_score:.2f}: {label} - {best.width}x{best.height} - {best.url}")
    else:
        summary.manual_required += 1
        mark_status(supabase_url, service_key, product["id"], "manual_required", args.dry_run)
        print(f"MANUAL {best.confidence_score:.2f}: {label} - best below threshold - {best.url}")

    save_candidates(supabase_url, service_key, product["id"], valid_candidates, approved_url, args.dry_run)


def main():
    parser = argparse.ArgumentParser(description="Official/retailer product image scraper for CrossCheckHealth.")
    parser.add_argument("--sheet", help="Path to CSV/XLSX product sheet. Defaults to newest .xlsx/.csv in project root.")
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only-missing", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--min-confidence", type=float, default=0.75)
    parser.add_argument("--delay", type=float, default=1.0)
    parser.add_argument("--source", choices=["official", "retailer", "all"], default="all")
    parser.add_argument("--max-candidates", type=int, default=10)
    args = parser.parse_args()

    load_env(ENV_FILE)
    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise SystemExit("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")

    sheet_path = detect_sheet(args.sheet)
    sheet_rows = read_sheet(sheet_path)
    by_id, by_brand_name = build_sheet_indexes(sheet_rows)
    print(f"Sheet: {sheet_path if sheet_path else 'not found'}")
    print(f"Sheet rows: {len(sheet_rows)}")
    ensure_bucket(supabase_url, service_key, args.dry_run)

    products = fetch_products(supabase_url, service_key, args.only_missing, args.force, args.limit)
    robots_cache = {}
    summary = Summary()
    for product in products:
        sheet_row = by_id.get(str(product.get("source_product_id") or "")) or by_id.get(str(product.get("id") or ""))
        if not sheet_row:
            sheet_row = by_brand_name.get(normalize_key(product.get("brand"), product.get("name")))
        try:
            process_product(product, sheet_row, args, supabase_url, service_key, robots_cache, summary)
        except (HTTPError, URLError, TimeoutError, RuntimeError) as error:
            summary.failed += 1
            print(f"FAILED unexpected: {product.get('brand')} {product.get('name')} - {error}")
        time.sleep(args.delay)

    summary.print()


if __name__ == "__main__":
    main()
