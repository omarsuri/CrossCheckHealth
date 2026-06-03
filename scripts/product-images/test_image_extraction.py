from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook


INPUT_FILE = Path("product_image_next_steps.xlsx")
OUTPUT_FILE = Path("image_extraction_test.xlsx")
SHEET_NAME = "product_image_next_steps"
ROW_LIMIT = 15
TIMEOUT_SECONDS = 20


@dataclass
class ImageCandidate:
    url: str
    width: int = 0
    height: int = 0
    score: int = 0
    source: str = "img"

    @property
    def area(self) -> int:
        return self.width * self.height


class ProductImageHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.og_images: list[str] = []
        self.img_candidates: list[dict[str, str]] = []
        self.json_ld_blocks: list[str] = []
        self._in_json_ld = False
        self._current_script: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attr = {key.lower(): value or "" for key, value in attrs}
        tag = tag.lower()

        if tag == "meta":
            prop = (attr.get("property") or attr.get("name") or "").lower()
            content = attr.get("content", "").strip()
            if content and prop in {"og:image", "og:image:url", "og:image:secure_url"}:
                self.og_images.append(content)

        if tag == "script":
            script_type = attr.get("type", "").lower()
            self._in_json_ld = "ld+json" in script_type
            self._current_script = []

        if tag == "img":
            self.img_candidates.append(attr)

    def handle_data(self, data: str) -> None:
        if self._in_json_ld:
            self._current_script.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "script" and self._in_json_ld:
            block = "".join(self._current_script).strip()
            if block:
                self.json_ld_blocks.append(block)
            self._in_json_ld = False
            self._current_script = []


def fetch_html(url: str) -> str:
    request = Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    with urlopen(request, timeout=TIMEOUT_SECONDS) as response:
        raw = response.read()
        content_type = response.headers.get("Content-Type", "")
        charset_match = re.search(r"charset=([^;\s]+)", content_type, re.I)
        encoding = charset_match.group(1) if charset_match else "utf-8"
        return raw.decode(encoding, errors="replace")


def is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def is_search_url(value: str) -> bool:
    parsed = urlparse(value)
    host = parsed.netloc.lower()
    full = value.lower()
    return (
        "google." in host
        or "bing." in host
        or "duckduckgo." in host
        or "yahoo." in host
        or "/search?" in full
        or "tbm=isch" in full
        or "images/search" in full
    )


def clean_image_url(base_url: str, value: str | None) -> str:
    if not value:
        return ""
    value = unescape(value).strip()
    if not value or value.startswith("data:"):
        return ""
    return urljoin(base_url, value)


def pick_from_srcset(base_url: str, srcset: str) -> ImageCandidate | None:
    best: ImageCandidate | None = None
    for part in srcset.split(","):
        pieces = part.strip().split()
        if not pieces:
            continue
        candidate_url = clean_image_url(base_url, pieces[0])
        if not candidate_url:
            continue
        width = 0
        if len(pieces) > 1:
            width_match = re.match(r"(\d+)w", pieces[1])
            density_match = re.match(r"([\d.]+)x", pieces[1])
            if width_match:
                width = int(width_match.group(1))
            elif density_match:
                width = int(float(density_match.group(1)) * 1000)
        candidate = ImageCandidate(url=candidate_url, width=width, score=width, source="srcset")
        if not best or candidate.score > best.score:
            best = candidate
    return best


def image_score(url: str, attrs: dict[str, str], width: int, height: int) -> int:
    lower = f"{url} {' '.join(attrs.values())}".lower()
    score = width * height if width and height else max(width, height, 0)
    for signal in ("product", "main", "primary", "gallery", "pdp", "sku", "pack", "front"):
        if signal in lower:
            score += 500_000
    for penalty in ("logo", "icon", "sprite", "placeholder", "loader", "avatar", "banner"):
        if penalty in lower:
            score -= 500_000
    if re.search(r"\.(jpg|jpeg|png|webp)(?:[?#]|$)", url, re.I):
        score += 100_000
    return score


def extract_json_ld_images(blocks: list[str], base_url: str) -> list[str]:
    images: list[str] = []

    def collect(value: Any) -> None:
        if isinstance(value, dict):
            if "image" in value:
                collect(value["image"])
            if "thumbnailUrl" in value:
                collect(value["thumbnailUrl"])
            if "url" in value and any(key in value for key in ("@type", "width", "height")):
                maybe_url = clean_image_url(base_url, str(value["url"]))
                if maybe_url:
                    images.append(maybe_url)
            for nested_key in ("@graph", "offers", "mainEntity", "itemListElement"):
                if nested_key in value:
                    collect(value[nested_key])
        elif isinstance(value, list):
            for item in value:
                collect(item)
        elif isinstance(value, str):
            maybe_url = clean_image_url(base_url, value)
            if maybe_url:
                images.append(maybe_url)

    for block in blocks:
        block = block.strip()
        if not block:
            continue
        try:
            collect(json.loads(block))
        except json.JSONDecodeError:
            # Some sites put several JSON objects in one script. Try a conservative object scan.
            for match in re.finditer(r"\{.*?\}", block, flags=re.S):
                try:
                    collect(json.loads(match.group(0)))
                except json.JSONDecodeError:
                    continue

    return dedupe(images)


def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def extract_largest_candidate(parser: ProductImageHTMLParser, base_url: str) -> str:
    candidates: list[ImageCandidate] = []
    for attrs in parser.img_candidates:
        srcset_candidate = pick_from_srcset(
            base_url,
            attrs.get("srcset") or attrs.get("data-srcset") or "",
        )
        if srcset_candidate:
            candidates.append(srcset_candidate)

        raw_url = (
            attrs.get("src")
            or attrs.get("data-src")
            or attrs.get("data-original")
            or attrs.get("data-lazy-src")
            or attrs.get("data-image")
            or ""
        )
        image_url = clean_image_url(base_url, raw_url)
        if not image_url:
            continue
        if image_url.lower().endswith(".svg") or image_url.lower().endswith(".gif"):
            continue
        width = int(attrs.get("width", "0") or 0) if (attrs.get("width", "0") or "0").isdigit() else 0
        height = int(attrs.get("height", "0") or 0) if (attrs.get("height", "0") or "0").isdigit() else 0
        candidates.append(
            ImageCandidate(
                url=image_url,
                width=width,
                height=height,
                score=image_score(image_url, attrs, width, height),
                source="img",
            )
        )

    if not candidates:
        return ""
    candidates.sort(key=lambda item: (item.score, item.area, item.width), reverse=True)
    return candidates[0].url


def extract_images(source_url: str) -> dict[str, str]:
    html = fetch_html(source_url)
    parser = ProductImageHTMLParser()
    parser.feed(html)

    og_image = clean_image_url(source_url, parser.og_images[0]) if parser.og_images else ""
    json_ld_images = extract_json_ld_images(parser.json_ld_blocks, source_url)
    json_ld_image = json_ld_images[0] if json_ld_images else ""
    largest_candidate = extract_largest_candidate(parser, source_url)
    image_url = og_image or json_ld_image or largest_candidate

    return {
        "og_image": og_image,
        "json_ld_image": json_ld_image,
        "largest_product_image_candidate": largest_candidate,
        "image_found": "yes" if image_url else "no",
        "image_url": image_url,
    }


def main() -> int:
    if not INPUT_FILE.exists():
        print(f"Input workbook not found: {INPUT_FILE}", file=sys.stderr)
        return 1

    input_wb = load_workbook(INPUT_FILE, data_only=True)
    if SHEET_NAME not in input_wb.sheetnames:
        print(f"Sheet not found: {SHEET_NAME}", file=sys.stderr)
        return 1

    input_ws = input_wb[SHEET_NAME]
    headers = {str(cell.value).strip(): index for index, cell in enumerate(input_ws[1], start=1)}
    for required in ("product_name", "source_url"):
        if required not in headers:
            print(f"Missing required column: {required}", file=sys.stderr)
            return 1

    results: list[dict[str, str]] = []
    for row in range(2, min(input_ws.max_row, ROW_LIMIT + 1) + 1):
        product_name = str(input_ws.cell(row, headers["product_name"]).value or "").strip()
        source_url = str(input_ws.cell(row, headers["source_url"]).value or "").strip()
        result = {
            "product_name": product_name,
            "source_url": source_url,
            "image_found": "no",
            "image_url": "",
            "og_image": "",
            "json_ld_image": "",
            "largest_product_image_candidate": "",
            "error": "",
        }

        if not source_url or not is_http_url(source_url) or is_search_url(source_url):
            result["error"] = "Missing, invalid, or search source_url"
        else:
            try:
                result.update(extract_images(source_url))
            except Exception as exc:  # noqa: BLE001 - keep test output complete for all 15 rows.
                result["error"] = f"{type(exc).__name__}: {exc}"

        print(f"{product_name}\t{source_url}\t{result['image_found']}\t{result['image_url']}")
        results.append(result)

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "image_extraction_test"
    output_headers = [
        "product_name",
        "source_url",
        "image_found",
        "image_url",
        "og_image",
        "json_ld_image",
        "largest_product_image_candidate",
        "error",
    ]
    output_ws.append(output_headers)
    for result in results:
        output_ws.append([result.get(header, "") for header in output_headers])

    for column_cells in output_ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        output_ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 80)

    output_wb.save(OUTPUT_FILE)
    print(f"Saved results to: {OUTPUT_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
